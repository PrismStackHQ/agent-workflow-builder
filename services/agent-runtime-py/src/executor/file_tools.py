"""Built-in file generation tools for the executor agent.

These tools generate PDF and Excel files locally. The agent passes structured
data (titles, headers, rows) and gets back a file path — no file content is
sent to the LLM, keeping token usage minimal.
"""

import json
import logging
import os
import uuid
from typing import Any

from langchain_core.tools import tool

logger = logging.getLogger(__name__)

# All generated files go here; mount as a Docker volume if persistence is needed
OUTPUT_DIR = os.environ.get("AGENT_OUTPUT_DIR", "/tmp/agent-outputs")


def _ensure_output_dir() -> str:
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    return OUTPUT_DIR


def _coerce_list(val: Any) -> list:
    """Coerce a value to a list — handles JSON strings from LLM tool calls."""
    if val is None:
        return []
    if isinstance(val, str):
        try:
            parsed = json.loads(val)
            if isinstance(parsed, list):
                return parsed
        except (json.JSONDecodeError, TypeError):
            pass
        return []
    if isinstance(val, list):
        return val
    return []


# ── PDF Generation ────────────────────────────────────────────────────────────


@tool
async def generate_pdf(
    title: str,
    content: str,
    table_headers: list[str] | None = None,
    table_rows: list[list[Any]] | None = None,
    filename: str | None = None,
) -> str:
    """Generate a PDF document with optional table data.

    Args:
        title: Document title displayed at the top.
        content: Body text / description paragraphs (plain text, newlines preserved).
        table_headers: Optional column headers for a data table.
        table_rows: Optional list of rows (each row is a list of cell values) for the table.
        filename: Optional output filename (without extension). Auto-generated if omitted.

    Returns:
        The absolute file path of the generated PDF.
    """
    try:
        from fpdf import FPDF

        out_dir = _ensure_output_dir()
        fname = (filename or f"report_{uuid.uuid4().hex[:8]}") + ".pdf"
        filepath = os.path.join(out_dir, fname)

        pdf = FPDF()
        pdf.set_auto_page_break(auto=True, margin=15)

        # Register DejaVu Sans for full Unicode support (Sinhala, CJK, etc.)
        font_dir = "/usr/share/fonts/truetype/dejavu"
        if os.path.isdir(font_dir):
            pdf.add_font("DejaVu", "", os.path.join(font_dir, "DejaVuSans.ttf"), uni=True)
            pdf.add_font("DejaVu", "B", os.path.join(font_dir, "DejaVuSans-Bold.ttf"), uni=True)
            font_family = "DejaVu"
        else:
            font_family = "Helvetica"

        pdf.add_page()

        # Title
        pdf.set_font(font_family, "B", 16)
        pdf.cell(0, 10, title, new_x="LMARGIN", new_y="NEXT", align="C")
        pdf.ln(5)

        # Body text
        if content:
            pdf.set_font(font_family, "", 11)
            pdf.multi_cell(0, 6, content)
            pdf.ln(5)

        # Table — coerce inputs
        table_headers = _coerce_list(table_headers)
        table_rows = _coerce_list(table_rows)
        if table_headers and table_rows:
            pdf.set_font(font_family, "B", 10)
            col_count = len(table_headers)
            page_width = pdf.w - pdf.l_margin - pdf.r_margin
            col_width = page_width / col_count

            # Header row
            for header in table_headers:
                pdf.cell(col_width, 8, str(header)[:30], border=1, align="C")
            pdf.ln()

            # Data rows
            pdf.set_font(font_family, "", 9)
            for row in table_rows:
                for i, cell in enumerate(row):
                    text = str(cell) if cell is not None else ""
                    pdf.cell(col_width, 7, text[:40], border=1)
                pdf.ln()

        pdf.output(filepath)
        logger.info(f"PDF generated: {filepath}")
        return f"PDF created successfully: {filepath} ({os.path.getsize(filepath)} bytes)"

    except ImportError:
        return "PDF generation not available (fpdf2 not installed)"
    except Exception as e:
        logger.error(f"PDF generation failed: {e}", exc_info=True)
        return f"PDF generation failed: {e}"


# ── Excel Generation ──────────────────────────────────────────────────────────


@tool
async def generate_excel(
    sheet_name: str,
    headers: list[str],
    rows: list[list[Any]],
    filename: str | None = None,
    title: str | None = None,
) -> str:
    """Generate an Excel (.xlsx) spreadsheet with structured data.

    Args:
        sheet_name: Name of the worksheet.
        headers: Column header names.
        rows: List of rows, each row is a list of cell values.
        filename: Optional output filename (without extension). Auto-generated if omitted.
        title: Optional title row at the top of the sheet.

    Returns:
        The absolute file path of the generated Excel file.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill

        # Coerce inputs — LLM may pass JSON strings instead of lists
        headers = _coerce_list(headers)
        rows = _coerce_list(rows)
        # Ensure each row is a list (LLM might pass list of dicts)
        coerced_rows = []
        for r in rows:
            if isinstance(r, dict):
                # Extract values in header order if possible, else dict values
                coerced_rows.append([r.get(h, r.get(h.lower(), "")) for h in headers] if headers else list(r.values()))
            elif isinstance(r, str):
                try:
                    parsed = json.loads(r)
                    if isinstance(parsed, list):
                        coerced_rows.append(parsed)
                    elif isinstance(parsed, dict):
                        coerced_rows.append([parsed.get(h, parsed.get(h.lower(), "")) for h in headers] if headers else list(parsed.values()))
                    else:
                        coerced_rows.append([parsed])
                except (json.JSONDecodeError, TypeError):
                    coerced_rows.append([r])
            elif isinstance(r, list):
                coerced_rows.append(r)
            else:
                coerced_rows.append([r])
        rows = coerced_rows

        logger.info(f"generate_excel called: sheet={sheet_name}, headers={headers}, rows_count={len(rows)}")
        if not headers:
            return "Excel generation failed: no headers provided. Pass headers as a list of column names."
        if not rows:
            return "Excel generation failed: no data rows provided. Pass rows as a list of lists."

        out_dir = _ensure_output_dir()
        fname = (filename or f"data_{uuid.uuid4().hex[:8]}") + ".xlsx"
        filepath = os.path.join(out_dir, fname)

        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name[:31]  # Excel sheet name limit

        start_row = 1

        # Optional title row
        if title:
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
            title_cell = ws.cell(row=1, column=1, value=title)
            title_cell.font = Font(bold=True, size=14)
            title_cell.alignment = Alignment(horizontal="center")
            start_row = 3  # Leave a blank row after title

        # Header row
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=start_row, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        # Data rows
        for row_idx, row_data in enumerate(rows, start_row + 1):
            for col_idx, value in enumerate(row_data, 1):
                # Convert any remaining dicts/lists to strings
                if isinstance(value, (dict, list)):
                    value = json.dumps(value, default=str)
                ws.cell(row=row_idx, column=col_idx, value=value)

        # Auto-fit column widths (approximate)
        from openpyxl.utils import get_column_letter
        for col_idx, header in enumerate(headers, 1):
            max_len = len(str(header))
            for row_data in rows[:50]:  # Sample first 50 rows
                if col_idx - 1 < len(row_data):
                    max_len = max(max_len, len(str(row_data[col_idx - 1] or "")))
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 3, 50)

        wb.save(filepath)
        logger.info(f"Excel generated: {filepath} ({len(rows)} rows, {len(headers)} cols)")
        return f"Excel file created successfully: {filepath} ({len(rows)} data rows, {len(headers)} columns, {os.path.getsize(filepath)} bytes)"

    except ImportError:
        return "Excel generation not available (openpyxl not installed)"
    except Exception as e:
        logger.error(f"Excel generation failed: {e}", exc_info=True)
        return f"Excel generation failed: {e}"


# ── CSV Generation ────────────────────────────────────────────────────────────


@tool
async def generate_csv(
    headers: list[str],
    rows: list[list[Any]],
    filename: str | None = None,
) -> str:
    """Generate a CSV file with structured data.

    Args:
        headers: Column header names.
        rows: List of rows, each row is a list of cell values.
        filename: Optional output filename (without extension). Auto-generated if omitted.

    Returns:
        The absolute file path of the generated CSV file.
    """
    try:
        import csv

        headers = _coerce_list(headers)
        rows = _coerce_list(rows)

        out_dir = _ensure_output_dir()
        fname = (filename or f"data_{uuid.uuid4().hex[:8]}") + ".csv"
        filepath = os.path.join(out_dir, fname)

        with open(filepath, "w", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)
            writer.writerow(headers)
            for r in rows:
                if isinstance(r, dict):
                    writer.writerow([r.get(h, "") for h in headers] if headers else r.values())
                elif isinstance(r, list):
                    writer.writerow(r)
                else:
                    writer.writerow([r])

        logger.info(f"CSV generated: {filepath}")
        return f"CSV file created successfully: {filepath} ({len(rows)} data rows, {os.path.getsize(filepath)} bytes)"

    except Exception as e:
        logger.error(f"CSV generation failed: {e}", exc_info=True)
        return f"CSV generation failed: {e}"


# ── Convenience export ────────────────────────────────────────────────────────

FILE_TOOLS = [generate_pdf, generate_excel, generate_csv]
